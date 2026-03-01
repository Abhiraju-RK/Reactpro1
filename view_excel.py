from openpyxl import load_workbook

wb=load_workbook("Demo.xlsx")
act=wb.active

for row in act.iter_rows(values_only=True):
    print(row)

